import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from django.conf import settings

from contrib.utils.excel.openpyxl import StreamToExcel, ExcelRendererBase
from ..services.prices import get_prices_list


ASSETS_DIR = settings.BASE_DIR / "apps/assets"


class PricesRenderer(ExcelRendererBase):
    def process(self):
        _ws = self._template.active
        _fonts = [Font(sz=size, name="Arial") for size in (14, 14, 12, 10)]
        _number_format = "# ##0"
        _center = Alignment(horizontal="center")

        for i, item in enumerate(self._source["catalog"], start=4):
            for j, key in enumerate((
                "title", "price", "quantity", "sold",
            ), start=1):
                _level = item["level"]
                _cell = _ws.cell(i, j, value=item[key] or "")
                _cell.font = _fonts[_level]
                _side = Side(border_style="dotted", color="888888")
                _cell.border = Border(left=_side, right=_side, top=_side, bottom=_side)
                if j > 1:
                    _cell.number_format = _number_format
                    _cell.alignment = _center
                if _level == 1:
                    _cell.fill = PatternFill("solid", fgColor="5f9ea0")

        return self._template


def get_prices_excel(request):
    _prices = get_prices_list(request, apply_intcomma=False)
    _template_path = ASSETS_DIR / "reports/prices.xlsx"
    return StreamToExcel(_prices, _template_path, PricesRenderer,
        before=[]).output

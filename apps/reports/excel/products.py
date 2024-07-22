from copy import copy

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from django.conf import settings

from contrib.utils.date_time import get_date_interval
from contrib.utils.excel.openpyxl import StreamToExcel, ExcelRendererBase
from ..services import get_products_report

REPORTS_DIR = settings.BASE_DIR / "apps/assets/reports"


class ProductsReportRenderer(ExcelRendererBase):
    def process(self):
        _ws = self._template.active
        _ws["L1"] = "за период: " + (
            get_date_interval(self._source["request"].GET) or "весь")
        _fonts = [Font(sz=size, name="Arial") for size in (14, 14, 12, 10)]
        _number_format = "# ##0"
        _center = Alignment(horizontal="center")

        for i, item in enumerate(self._source["totals"], start=5):
            for j, key in enumerate((
                "title", "quantity_start", "supplies_amount", "supplies_quantity",
                "sales_amount", "sales_quantity", "sales_scrap_weight",
                "margin", "margin_percent", "margin_with_scrap", "amount_end", "quantity_end"
            ), start=1):
                _level = item["level"]
                _cell = _ws.cell(i, j, value=item[key] or "")
                _cell.font = _fonts[_level]
                _side = Side(border_style="dotted", color="888888")
                _cell.border = Border(left=_side, right=_side, top=_side, bottom=_side)
                if j > 1:
                    _cell.number_format = _number_format
                    _cell.alignment = _center
                if _level == 1:
                    _cell.fill = PatternFill("solid", fgColor="5f9ea0")
                if key in ("sales_scrap_weight", "margin_percent", "margin_with_scrap"):
                    _font = copy(_cell.font)
                    _font.italic = True
                    _cell.font = _font

        return self._template


def get_products_report_excel(request):
    _totals = get_products_report(request, apply_intcomma=False)
    _template_path = REPORTS_DIR / "stock.xlsx"
    return StreamToExcel(_totals, _template_path, ProductsReportRenderer).output
